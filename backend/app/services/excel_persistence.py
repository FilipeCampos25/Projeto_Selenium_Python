"""
excel_persistence.py
Gerencia a persistência de dados no arquivo Excel.
MODIFICADO PARA EXECUÇÃO LOCAL - Salva na pasta raiz do projeto.
"""
import os
import logging
from typing import List, Dict, Any
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)

class ExcelPersistence:
    """
    Gerencia a persistência de dados no arquivo Excel.
    MODIFICADO PARA EXECUÇÃO LOCAL.
    """
    def __init__(self, file_path: str = None):
        # ============================================================
        # 🔴 INÍCIO MODIFICAÇÃO LOCAL - REMOVER QUANDO VOLTAR DOCKER
        # ============================================================
        
        # Se não especificar caminho, salva na pasta raiz do projeto
        if file_path is None:
            # Salva na pasta outputs_local na raiz
            outputs_dir = os.path.join(os.getcwd(), "outputs_local")
            os.makedirs(outputs_dir, exist_ok=True)
            file_path = os.path.join(outputs_dir, "PGC_2025.xlsx")
            logger.info(f"[LOCAL] Excel será salvo em: {outputs_dir}")
        
        self.file_path = file_path
        self._ensure_file_exists()
        
        # ============================================================
        # 🔴 FIM MODIFICAÇÃO LOCAL
        # ============================================================
        
        # CÓDIGO ORIGINAL (quando voltar Docker, ajustar paths):
        # self.file_path = file_path
        # self._ensure_file_exists()

    def _ensure_file_exists(self):
        """Garante que o arquivo Excel exista, criando-o se necessário."""
        if not os.path.exists(self.file_path):
            logger.info(f"[LOG-VBA] Criando novo arquivo Excel: {self.file_path}")
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb["Sheet"].title = "PGC"
            else:
                wb.create_sheet("PGC")
            
            wb.create_sheet("PNCP")
            wb.create_sheet("Geral")
            
            # Cabeçalhos PGC
            ws_pgc = wb["PGC"]
            headers_pgc = ["Pag", "DFD", "Requisitante", "Descrição", "Valor", 
                          "Situação", "Conclusão", "Editor", "Responsáveis", "PTA", "Justificativa"]
            for col, header in enumerate(headers_pgc, 1):
                ws_pgc.cell(row=1, column=col, value=header)
            
            # Cabeçalhos PNCP
            ws_pncp = wb["PNCP"]
            headers_pncp = ["Contratação", "Descrição", "Categoria", "Valor", "Início", 
                           "Fim", "Status", "PGC", "DFD", "Status", "Tipo"]
            for col, header in enumerate(headers_pncp, 1):
                ws_pncp.cell(row=1, column=col, value=header)

            wb.save(self.file_path)
            logger.info(f"[LOCAL] ✅ Arquivo Excel criado: {self.file_path}")

    def update_pgc_sheet(self, data: List[Dict[str, Any]]):
        """
        Atualiza a aba PGC seguindo a lógica do VBA.
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb["PGC"]
            
            for item in data:
                dfd = str(item.get("dfd", "")).strip()
                if not dfd: 
                    continue
                    
                # Procura se DFD já existe
                found_row = None
                for row in range(2, ws.max_row + 1):
                    if str(ws.cell(row=row, column=2).value).strip() == dfd:
                        found_row = row
                        break
                
                # Se não existe, adiciona nova linha
                target_row = found_row if found_row else ws.max_row + 1
                
                # Preenche dados
                ws.cell(row=target_row, column=1, value=item.get("pag", 1))
                ws.cell(row=target_row, column=2, value=dfd)
                ws.cell(row=target_row, column=3, value=item.get("requisitante"))
                ws.cell(row=target_row, column=4, value=item.get("descricao"))
                ws.cell(row=target_row, column=5, value=item.get("valor"))
                ws.cell(row=target_row, column=6, value=item.get("situacao"))
                ws.cell(row=target_row, column=7, value=item.get("conclusao"))
                ws.cell(row=target_row, column=8, value=item.get("editor"))
                ws.cell(row=target_row, column=9, value=item.get("responsaveis"))
                ws.cell(row=target_row, column=10, value=item.get("pta"))
                ws.cell(row=target_row, column=11, value=item.get("justificativa"))
            
            wb.save(self.file_path)
            logger.info(f"[LOCAL] ✅ Aba PGC atualizada ({len(data)} itens)")
            
        except Exception as e:
            logger.error(f"[LOCAL] ❌ Erro ao atualizar aba PGC: {e}")

    def update_pncp_sheet(self, data: List[Dict[str, Any]]):
        """
        Atualiza a aba PNCP seguindo fielmente o mapeamento de colunas do VBA (A a K).
        """
        logger.info(f"[LOG-VBA] Iniciando atualização da aba PNCP no Excel...")
        try:
            wb = load_workbook(self.file_path)
            
            if "PNCP" not in wb.sheetnames:
                wb.create_sheet("PNCP")
            
            ws = wb["PNCP"]
            
            # Limpar dados antigos
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)

            # Estilização do cabeçalho
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True)
            for col in range(1, 12):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, entry in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=entry.get("col_a_contratacao"))
                ws.cell(row=row_idx, column=2, value=entry.get("col_b_descricao"))
                ws.cell(row=row_idx, column=3, value=entry.get("col_c_categoria"))
                
                cell_valor = ws.cell(row=row_idx, column=4, value=entry.get("col_d_valor"))
                cell_valor.number_format = '"R$" #,##0.00'
                
                ws.cell(row=row_idx, column=5, value=entry.get("col_e_inicio"))
                ws.cell(row=row_idx, column=6, value=entry.get("col_f_fim"))
                ws.cell(row=row_idx, column=7, value=entry.get("col_g_status"))
                ws.cell(row=row_idx, column=8, value=entry.get("col_h_status_tipo"))
                
                cell_dfd = ws.cell(row=row_idx, column=9, value=entry.get("col_i_dfd"))
                cell_dfd.number_format = '@'
                
                ws.cell(row=row_idx, column=10, value=entry.get("col_g_status"))
                ws.cell(row=row_idx, column=11, value=entry.get("col_h_status_tipo"))

            # Ajuste automático de largura
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: 
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            wb.save(self.file_path)
            logger.info(f"[LOCAL] ✅ Aba PNCP atualizada ({len(data)} itens)")
            
        except Exception as e:
            logger.error(f"[LOCAL] ❌ Erro ao atualizar aba PNCP: {e}")

    def sync_to_geral(self):
        """
        Sincroniza dados entre PGC e Geral seguindo a lógica do VBA.
        """
        try:
            wb = load_workbook(self.file_path)
            if "PGC" not in wb.sheetnames or "Geral" not in wb.sheetnames:
                return
                
            ws_pgc = wb["PGC"]
            ws_geral = wb["Geral"]
            
            for r_pgc in range(2, ws_pgc.max_row + 1):
                dfd = ws_pgc.cell(row=r_pgc, column=2).value
                if not dfd: 
                    continue
                    
                found_row_geral = None
                for r_geral in range(2, ws_geral.max_row + 1):
                    if ws_geral.cell(row=r_geral, column=7).value == dfd:
                        found_row_geral = r_geral
                        break
                        
                target_row = found_row_geral if found_row_geral else ws_geral.max_row + 1
                ws_geral.cell(row=target_row, column=1, value=ws_pgc.cell(row=r_pgc, column=1).value)
                ws_geral.cell(row=target_row, column=6, value=str(dfd)[-4:] if dfd else "")
                ws_geral.cell(row=target_row, column=7, value=dfd)
                ws_geral.cell(row=target_row, column=8, value=ws_pgc.cell(row=r_pgc, column=3).value)
                ws_geral.cell(row=target_row, column=9, value=ws_pgc.cell(row=r_pgc, column=4).value)
                ws_geral.cell(row=target_row, column=10, value=ws_pgc.cell(row=r_pgc, column=5).value)
                ws_geral.cell(row=target_row, column=11, value=ws_pgc.cell(row=r_pgc, column=6).value)
                
            wb.save(self.file_path)
            logger.info("[LOCAL] ✅ Sincronização com aba Geral concluída")
            
        except Exception as e:
            logger.error(f"[LOCAL] ❌ Erro na sincronização Geral: {e}")